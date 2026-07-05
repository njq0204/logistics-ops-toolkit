"""Excel 专业报表导出 — 多 Sheet + 样式 + 条件格式 + 图表嵌入"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Union

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# 列名中文化映射
COLUMN_CN = {
    "region": "区域",
    "order_count": "单量",
    "total_revenue": "营收(元)",
    "total_cost": "成本(元)",
    "avg_weight": "均重(kg)",
    "avg_satisfaction": "满意度",
    "avg_delivery": "时效(天)",
    "gross_profit": "毛利(元)",
    "margin_rate": "毛利率(%)",
    "month": "月份",
    "revenue_mom(%)": "营收环比(%)",
    "order_mom(%)": "单量环比(%)",
    "profit_mom(%)": "利润环比(%)",
    "product_type": "产品类型",
    "revenue_share": "营收占比(%)",
    "customer_type": "客户类型",
    "avg_order_value": "客单价(元)",
    "return_count": "退单数",
    "return_rate(%)": "退单率(%)",
    "customer_name": "客户名称",
    "sales_rep": "销售员",
    "customer_count": "客户数",
    "city": "城市",
    "month_num": "月份",
    "revenue_yoy(%)": "营收同比(%)",
    "order_yoy(%)": "单量同比(%)",
    "name": "指标",
    "target": "目标值",
    "actual": "实际值",
    "progress_pct": "完成率(%)",
    "gap": "差距",
    "status": "状态码",
    "status_label": "状态",
    "period": "周期",
    "abc_class": "ABC分类",
    "distance_type": "距离类型",
    "revenue_per_kg": "单公斤营收(元)",
    "profit_per_kg": "单公斤毛利(元)",
    "total_weight": "计费重量(kg)",
    "value": "数值",
    "level": "级别",
    "category": "类别",
    "message": "预警内容",
    "priority": "优先级",
    "area": "领域",
    "action": "行动建议",
    "expected_impact": "预期效果",
}

# 样式常量
HEADER_FILL = PatternFill("solid", fgColor="0066CC")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=16, color="0066CC")
SUBTITLE_FONT = Font(name="微软雅黑", bold=True, size=12, color="333333")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")


def _rename_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [COLUMN_CN.get(c, c) for c in out.columns]
    return out


def _auto_width(ws, min_width=10, max_width=36):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 3, min_width), max_width)


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_area(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def _write_df(ws, df: pd.DataFrame, start_row=1, rename=True):
    data = _rename_columns(df) if rename else df
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    _style_header_row(ws, start_row)
    _style_data_area(ws, start_row + 1)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    _auto_width(ws)
    return start_row + len(data) + 1


def _apply_margin_format(ws, col_name="毛利率(%)"):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if col_name not in headers:
        return
    col_idx = headers.index(col_name) + 1
    letter = get_column_letter(col_idx)
    rng = f"{letter}2:{letter}{ws.max_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["20"], fill=RED_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["28"], fill=GREEN_FILL))


def _apply_goal_format(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "完成率(%)" not in headers:
        return
    col_idx = headers.index("完成率(%)") + 1
    letter = get_column_letter(col_idx)
    rng = f"{letter}2:{letter}{ws.max_row}"
    ws.conditional_formatting.add(
        rng,
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100,
                    color="0066CC", showValue=True, minLength=0, maxLength=100),
    )
    if "状态" in headers:
        status_col = get_column_letter(headers.index("状态") + 1)
        for r in range(2, ws.max_row + 1):
            val = ws[f"{status_col}{r}"].value
            row_fill = GREEN_FILL if val == "已达标" else RED_FILL if val == "落后" else YELLOW_FILL if val == "有风险" else None
            if row_fill:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = row_fill


def _apply_alert_format(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "级别" not in headers:
        return
    col = get_column_letter(headers.index("级别") + 1)
    for r in range(2, ws.max_row + 1):
        level = ws[f"{col}{r}"].value
        fill = RED_FILL if level == "danger" else YELLOW_FILL if level == "warning" else None
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill


def _build_overview_sheet(ws, project_name, run_at, data_summary, metrics, goal_summary, alerts, diagnosis=None):
    ws.sheet_properties.tabColor = "0066CC"
    ws.merge_cells("A1:F1")
    ws["A1"] = project_name
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"经营分析报告  |  生成时间: {run_at}"
    ws["A2"].font = Font(name="微软雅黑", size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    ws.cell(row, 1, "数据概览").font = SUBTITLE_FONT
    row += 1
    for k, v in data_summary.items():
        ws.cell(row, 1, k).font = Font(bold=True)
        ws.cell(row, 2, v)
        row += 1

    row += 1
    ws.cell(row, 1, "核心 KPI").font = SUBTITLE_FONT
    row += 1
    kpi_headers = ["指标", "数值"]
    for c, h in enumerate(kpi_headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1
    kpi_start = row
    for k, v in metrics.items():
        ws.cell(row, 1, k)
        ws.cell(row, 2, v)
        row += 1
    _style_data_area(ws, kpi_start)

    row += 1
    ws.cell(row, 1, "目标达成汇总").font = SUBTITLE_FONT
    row += 1
    if goal_summary:
        for k, v in goal_summary.items():
            label = {"total": "总目标数", "achieved": "已达标", "on_track": "正常",
                     "at_risk": "有风险", "behind": "落后", "avg_progress": "平均完成率(%)"}.get(k, k)
            ws.cell(row, 1, label)
            ws.cell(row, 2, v)
            row += 1

    row += 1
    ws.cell(row, 1, "预警统计").font = SUBTITLE_FONT
    row += 1
    danger = sum(1 for a in alerts if a.get("level") == "danger")
    warning = sum(1 for a in alerts if a.get("level") == "warning")
    ws.cell(row, 1, "严重预警")
    ws.cell(row, 2, danger)
    ws.cell(row + 1, 1, "一般预警")
    ws.cell(row + 1, 2, warning)

    if diagnosis:
        row += 3
        ws.cell(row, 1, "经营诊断").font = SUBTITLE_FONT
        row += 1
        ws.cell(row, 1, diagnosis.get("summary", ""))
        row += 1
        ws.cell(row, 1, f"综合评分: {diagnosis.get('score', '-')}/100")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28


def _add_region_chart(wb, sheet_name="区域分析"):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    if ws.max_row < 3:
        return
    chart = BarChart()
    chart.type = "col"
    chart.title = "各区域营收对比"
    chart.y_axis.title = "营收(元)"
    chart.x_axis.title = "区域"
    chart.style = 10
    chart.width = 18
    chart.height = 10

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "区域" not in headers or "营收(元)" not in headers:
        return
    reg_col = headers.index("区域") + 1
    rev_col = headers.index("营收(元)") + 1
    cats = Reference(ws, min_col=reg_col, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=rev_col, min_row=1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{get_column_letter(ws.max_column + 2)}2")


def _add_trend_chart(wb, sheet_name="月度趋势"):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    if ws.max_row < 3:
        return
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "月份" not in headers or "营收(元)" not in headers:
        return

    chart = LineChart()
    chart.title = "月度营收趋势"
    chart.y_axis.title = "营收(元)"
    chart.style = 10
    chart.width = 20
    chart.height = 10

    month_col = headers.index("月份") + 1
    rev_col = headers.index("营收(元)") + 1
    cats = Reference(ws, min_col=month_col, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=rev_col, min_row=1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{get_column_letter(ws.max_column + 2)}2")


def _embed_chart_images(wb, chart_paths: List[Path]):
    if not chart_paths:
        return
    ws = wb.create_sheet("可视化图表", len(wb.sheetnames))
    ws.sheet_properties.tabColor = "FF9900"
    ws["A1"] = "分析图表汇总"
    ws["A1"].font = SUBTITLE_FONT

    row = 3
    col_offset = 0
    try:
        from openpyxl.drawing.image import Image
    except ImportError:
        ws["A3"] = "（图片模块不可用）"
        return

    for i, cp in enumerate(chart_paths[:6]):
        if not cp.exists():
            continue
        img = Image(str(cp))
        img.width = min(img.width, 480)
        img.height = min(img.height, 280)
        anchor_col = "A" if i % 2 == 0 else "H"
        anchor_row = row if i % 2 == 0 else row
        ws.add_image(img, f"{anchor_col}{anchor_row}")
        ws.cell(anchor_row - 1 if anchor_row > 1 else 1, 1 if i % 2 == 0 else 8, cp.stem)
        if i % 2 == 1:
            row += 16


def export_to_excel(
    output_path: Union[str, Path],
    metrics: dict,
    region_summary: pd.DataFrame,
    monthly: pd.DataFrame,
    product_mix: pd.DataFrame,
    customer_seg: pd.DataFrame,
    top_customers: pd.DataFrame,
    sales_ranking: pd.DataFrame,
    alerts: list,
    goals: list = None,
    goal_summary: dict = None,
    data_summary: dict = None,
    project_name: str = "物流经营数据分析平台",
    run_at: str = "",
    mom: pd.DataFrame = None,
    yoy: pd.DataFrame = None,
    city_ranking: pd.DataFrame = None,
    region_product: pd.DataFrame = None,
    chart_paths: List[Path] = None,
    raw_df: pd.DataFrame = None,
    include_raw: bool = False,
    max_raw_rows: int = 500,
    diagnosis: dict = None,
    customer_abc: pd.DataFrame = None,
    distance_mix: pd.DataFrame = None,
    unit_economics: pd.DataFrame = None,
) -> Path:
    """导出专业多 Sheet Excel 经营分析报告"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_at = run_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    metrics_df = pd.DataFrame(list(metrics.items()), columns=["指标", "数值"])
    alerts_df = pd.DataFrame(alerts) if alerts else pd.DataFrame([{"message": "暂无预警"}])
    goals_df = pd.DataFrame(goals) if goals else pd.DataFrame([{"message": "未配置目标"}])
    goal_summary = goal_summary or {}
    data_summary = data_summary or {}

    diagnosis = diagnosis or {}
    actions_df = pd.DataFrame(diagnosis.get("actions", []))
    if actions_df.empty:
        actions_df = pd.DataFrame([{"action": "暂无建议，经营状态良好"}])

    sheet_colors = {
        "报告总览": "0066CC", "经营诊断": "00AA66", "行动建议": "FF6600",
        "区域分析": "4472C4", "月度趋势": "4472C4", "环比分析": "5B9BD5",
        "同比分析": "5B9BD5", "产品结构": "ED7D31", "客户分层": "ED7D31",
        "TOP客户": "FFC000", "销售排名": "FFC000", "城市排名": "A5A5A5",
        "区域产品交叉": "A5A5A5", "经营预警": "FF0000", "可视化图表": "FF9900",
        "运单明细": "808080",
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 占位写入，后续用 openpyxl 精修
        pd.DataFrame().to_excel(writer, sheet_name="报告总览")
        metrics_df.to_excel(writer, sheet_name="核心KPI", index=False)
        goals_df.to_excel(writer, sheet_name="目标追踪", index=False)
        region_summary.to_excel(writer, sheet_name="区域分析", index=False)
        monthly.to_excel(writer, sheet_name="月度趋势", index=False)
        product_mix.to_excel(writer, sheet_name="产品结构", index=False)
        customer_seg.to_excel(writer, sheet_name="客户分层", index=False)
        top_customers.to_excel(writer, sheet_name="TOP客户", index=False)
        sales_ranking.to_excel(writer, sheet_name="销售排名", index=False)
        alerts_df.to_excel(writer, sheet_name="经营预警", index=False)
        actions_df.to_excel(writer, sheet_name="行动建议", index=False)

        if customer_abc is not None and not customer_abc.empty:
            customer_abc.to_excel(writer, sheet_name="ABC客户", index=False)
        if distance_mix is not None and not distance_mix.empty:
            distance_mix.to_excel(writer, sheet_name="距离结构", index=False)
        if unit_economics is not None and not unit_economics.empty:
            unit_economics.to_excel(writer, sheet_name="单公斤economics", index=False)

        if yoy is not None and not yoy.empty:
            yoy.to_excel(writer, sheet_name="同比分析", index=False)
        if city_ranking is not None and not city_ranking.empty:
            city_ranking.to_excel(writer, sheet_name="城市排名", index=False)
        if region_product is not None and not region_product.empty:
            region_product.to_excel(writer, sheet_name="区域产品交叉", index=False)
        if include_raw and raw_df is not None and not raw_df.empty:
            cols = [c for c in raw_df.columns if c not in ("gross_profit", "margin_rate", "month", "year", "year_month")]
            raw_df[cols].head(max_raw_rows).to_excel(writer, sheet_name="运单明细", index=False)

        wb = writer.book

        # 报告总览
        ws_overview = wb["报告总览"]
        _build_overview_sheet(ws_overview, project_name, run_at, data_summary, metrics, goal_summary, alerts, diagnosis)

        # 经营诊断 Sheet
        ws_diag = wb.create_sheet("经营诊断", 1)
        ws_diag.sheet_properties.tabColor = "00AA66"
        row = 1
        ws_diag.cell(row, 1, "经营诊断报告").font = SUBTITLE_FONT
        row += 1
        ws_diag.cell(row, 1, diagnosis.get("summary", ""))
        row += 2
        for title, key in [("亮点", "highlights"), ("风险", "risks"), ("机会", "opportunities")]:
            ws_diag.cell(row, 1, title).font = Font(bold=True, color="0066CC")
            row += 1
            for item in diagnosis.get(key, []):
                ws_diag.cell(row, 1, f"• {item}")
                row += 1
            row += 1

        df_map = {
            "核心KPI": metrics_df,
            "目标追踪": goals_df,
            "行动建议": actions_df,
            "区域分析": region_summary,
            "月度趋势": monthly,
            "产品结构": product_mix,
            "客户分层": customer_seg,
            "TOP客户": top_customers,
            "销售排名": sales_ranking,
            "经营预警": alerts_df,
        }
        if yoy is not None and not yoy.empty:
            df_map["同比分析"] = yoy
        if city_ranking is not None and not city_ranking.empty:
            df_map["城市排名"] = city_ranking
        if region_product is not None and not region_product.empty:
            df_map["区域产品交叉"] = region_product
        if customer_abc is not None and not customer_abc.empty:
            df_map["ABC客户"] = customer_abc
        if distance_mix is not None and not distance_mix.empty:
            df_map["距离结构"] = distance_mix
        if unit_economics is not None and not unit_economics.empty:
            df_map["单公斤economics"] = unit_economics
        if include_raw and raw_df is not None:
            cols = [c for c in raw_df.columns if c not in ("gross_profit", "margin_rate", "month", "year", "year_month")]
            df_map["运单明细"] = raw_df[cols].head(max_raw_rows)

        for sheet_name, df in df_map.items():
            if sheet_name == "报告总览":
                continue
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
            _write_df(ws, df, start_row=1)
            if sheet_name in sheet_colors:
                ws.sheet_properties.tabColor = sheet_colors[sheet_name]
            if sheet_name == "区域分析":
                _apply_margin_format(ws)
            if sheet_name == "目标追踪":
                _apply_goal_format(ws)
            if sheet_name == "经营预警":
                _apply_alert_format(ws)

        # 环比 Sheet
        mom_df = mom if mom is not None and not mom.empty else monthly
        mom_cols = [c for c in mom_df.columns if "mom" in c.lower() or c in ("month", "total_revenue", "order_count", "gross_profit", "margin_rate")]
        mom_show = mom_df[[c for c in mom_cols if c in mom_df.columns]]
        if "环比分析" in wb.sheetnames:
            del wb["环比分析"]
        ws_mom = wb.create_sheet("环比分析", 5)
        _write_df(ws_mom, mom_show)
        ws_mom.sheet_properties.tabColor = "5B9BD5"

        # 嵌入原生 Excel 图表
        _add_region_chart(wb, "区域分析")
        _add_trend_chart(wb, "月度趋势")
        _embed_chart_images(wb, chart_paths or [])

        # 调整 Sheet 顺序
        order = ["报告总览", "经营诊断", "行动建议", "核心KPI", "目标追踪", "区域分析", "单公斤economics",
                 "月度趋势", "环比分析", "同比分析", "产品结构", "距离结构", "客户分层", "ABC客户",
                 "TOP客户", "销售排名", "城市排名", "区域产品交叉", "经营预警", "可视化图表", "运单明细"]
        for i, name in enumerate(order):
            if name in wb.sheetnames:
                wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    return output_path
